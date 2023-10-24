import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

# Configurações iniciais
FILENAME = "cadastro.xlsx"
if not os.path.exists(FILENAME):
    # Cria uma planilha do zero se não existir
    wb = Workbook()
    ws = wb.active
    ws.append(["Tipo", "Nome", "Telefone", "CPF/CNPJ", "E-mail"])
    wb.save(FILENAME)


def authenticate(username, password):
    # Função de autenticação simples (não é seguro para produção!)
    if username == "LOGIN DEFINIDA PELO USUÁRIO" and password == "SENHA DEFINIDA PELO USUÁRIO ":
        return True
    return False


def save_data(tipo, nome, telefone, cpf_cnpj, email):
    # Salva os dados na planilha
    wb = load_workbook(FILENAME)
    ws = wb.active
    ws.append([tipo, nome, telefone, cpf_cnpj, email])
    wb.save(FILENAME)


def get_data():
    # Busca todos os dados da planilha
    wb = load_workbook(FILENAME)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))[1:]


def login_screen():
    # Tela de login
    root = tk.Tk()
    root.title("Sistema de Gerenciamento de Clientes")
    root.geometry("400x200")

    def on_submit():
        if authenticate(username_entry.get(), password_entry.get()):
            root.destroy()
            main_screen()
        else:
            messagebox.showerror("Erro", "Credenciais inválidas!")

    frame = ttk.Frame(root, padding="30")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # Configuração para expandir corretamente
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(0, weight=1)

    frame.grid_columnconfigure(1, weight=1)

    # Widgets para entrada de nome de usuário e senha
    ttk.Label(frame, text="Nome de usuário:").grid(
        row=0, column=0, sticky=tk.E, pady=10
    )
    username_entry = ttk.Entry(frame)
    username_entry.grid(row=0, column=1, sticky="ew", pady=10)

    ttk.Label(frame, text="Senha:").grid(row=1, column=0, sticky=tk.E, pady=10)
    password_entry = ttk.Entry(frame, show="*")
    password_entry.grid(row=1, column=1, sticky="ew", pady=10)

    ttk.Button(frame, text="Entrar", command=on_submit).grid(
        row=2, column=0, columnspan=2, pady=20, sticky="ew"
    )

    root.mainloop()


def main_screen():
    # Tela principal de cadastro
    root = tk.Tk()
    root.title("Cadastro")
    root.geometry("600x300")

    def on_submit():
        tipo = "PF" if tipo_var.get() == 1 else "PJ"
        save_data(
            tipo,
            nome_entry.get(),
            telefone_entry.get(),
            cpf_entry.get(),
            email_entry.get(),
        )
        nome_entry.delete(0, tk.END)
        telefone_entry.delete(0, tk.END)
        cpf_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)

    def on_query():
        root.destroy()
        query_screen()

    frame = ttk.Frame(root, padding="30")
    frame.pack(pady=30)

    # Widgets para entrada de dados
    tipo_var = tk.IntVar()
    tipo_var.set(1)
    ttk.Radiobutton(frame, text="Pessoa Física", variable=tipo_var, value=1).grid(
        row=0, column=1, sticky=tk.W, pady=5
    )
    ttk.Radiobutton(frame, text="Pessoa Jurídica", variable=tipo_var, value=2).grid(
        row=0, column=2, sticky=tk.W, pady=5
    )

    # Centralizando os campos de entrada
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_columnconfigure(3, weight=1)

    ttk.Label(frame, text="Nome:").grid(row=1, column=1, sticky=tk.E, pady=5)
    nome_entry = ttk.Entry(frame, width=30)
    nome_entry.grid(row=1, column=2, sticky="ew", pady=5)

    ttk.Label(frame, text="Telefone:").grid(row=2, column=1, sticky=tk.E, pady=5)
    telefone_entry = ttk.Entry(frame, width=30)
    telefone_entry.grid(row=2, column=2, sticky="ew", pady=5)

    ttk.Label(frame, text="CPF/CNPJ:").grid(row=3, column=1, sticky=tk.E, pady=5)
    cpf_entry = ttk.Entry(frame, width=30)
    cpf_entry.grid(row=3, column=2, sticky="ew", pady=5)

    ttk.Label(frame, text="E-mail:").grid(row=4, column=1, sticky=tk.E, pady=5)
    email_entry = ttk.Entry(frame, width=30)
    email_entry.grid(row=4, column=2, sticky="ew", pady=5)

    ttk.Button(frame, text="Salvar", command=on_submit).grid(row=5, column=1, pady=20)
    ttk.Button(frame, text="Consultar Dados", command=on_query).grid(
        row=5, column=2, pady=20
    )

    root.mainloop()


def query_screen():
    # Tela de consulta
    root = tk.Tk()
    root.title("Consulta")
    root.geometry("700x400")

    frame = ttk.Frame(root, padding="30")
    frame.pack(pady=30)

    # Cria um treeview para apresentar os dados em formato de tabela
    columns = ("Tipo", "Nome", "Telefone", "CPF/CNPJ", "E-mail")
    tree = ttk.Treeview(frame, columns=columns, show="headings")

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=120, stretch=tk.YES)

    tree.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Carrega os dados na treeview
    data = get_data()
    for row in data:
        tree.insert("", tk.END, values=row)

    # Adiciona um botão de retorno
    ttk.Button(frame, text="Voltar", command=root.destroy).pack(pady=20)

    root.mainloop()


# Iniciar o sistema com a tela de login
login_screen()
